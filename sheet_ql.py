"""
SheetQL: Professional Data Analysis & ETL Tool

This module implements an interactive Command Line Interface (CLI) for querying
flat files (CSV, Excel, Parquet, JSON) using SQL. It leverages DuckDB for
high-performance in-memory processing and provides a "Zero-Copy" architecture
for handling large datasets efficiently.
"""

import os
import re
import argparse
import logging
import warnings
from collections import deque
from typing import Any, Optional, List, Tuple, Dict

import pandas as pd
import duckdb
from rich.console import Console
from rich.logging import RichHandler
from rich.table import Table
from openpyxl.styles import Font, PatternFill

try:
    import python_calamine

    CALAMINE_AVAILABLE = True
except ImportError:
    CALAMINE_AVAILABLE = False

try:
    import xlsxwriter

    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False

try:
    from prompt_toolkit import PromptSession
    from prompt_toolkit.completion import Completer, Completion
    from prompt_toolkit.lexers import PygmentsLexer
    from pygments.lexers.sql import SqlLexer
    from prompt_toolkit.styles import Style

    PROMPT_TOOLKIT_AVAILABLE = True
except ImportError:
    PROMPT_TOOLKIT_AVAILABLE = False

try:
    import yaml

    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

try:
    import tkinter as tk
    from tkinter import filedialog

    TKINTER_AVAILABLE = True
except ImportError:
    TKINTER_AVAILABLE = False


def setup_logging(debug_mode: bool = False) -> logging.Logger:
    """
    Configures the application logging system with File and Console handlers.
    """
    logger = logging.getLogger("SheetQL")
    logger.setLevel(logging.DEBUG)
    logger.handlers = []

    file_handler = logging.FileHandler("sheetql.log", mode="w", encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_formatter = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s"
    )
    file_handler.setFormatter(file_formatter)
    logger.addHandler(file_handler)

    console_level = logging.DEBUG if debug_mode else logging.INFO
    rich_handler = RichHandler(
        console=Console(stderr=True), show_time=False, show_path=False, markup=True
    )
    rich_handler.setLevel(console_level)
    logger.addHandler(rich_handler)

    for lib in ["duckdb", "matplotlib", "PIL", "prompt_toolkit"]:
        lib_logger = logging.getLogger(lib)
        lib_logger.setLevel(logging.WARNING)
        lib_logger.propagate = True

    logging.captureWarnings(True)
    return logger


class SessionRecorder:
    """
    Records session activities (Loads, Queries, Exports) to generate YAML scripts.
    """

    def __init__(self):
        self.inputs: List[Dict[str, str]] = []
        self.transformations: List[Dict[str, str]] = []
        self.exports: List[Dict[str, str]] = []

    def record_load(self, path: str, alias: str) -> None:
        """Records a file load operation."""
        self.inputs.append({"path": path, "alias": alias})

    def record_query(self, name: str, sql: str) -> None:
        """Records a transformation query, ignoring metadata commands."""
        if sql.strip().upper().startswith(("SHOW", "DESCRIBE", "PRAGMA")):
            return
        self.transformations.append({"name": name, "sql": sql})

    def record_export(self, path: str) -> None:
        """Records an export operation."""
        self.exports.append({"path": path})

    def generate_yaml(self) -> str:
        """Serializes the recorded session into a YAML string."""
        if not YAML_AVAILABLE:
            return "# Error: PyYAML not installed."

        script = {}
        if self.inputs:
            script["inputs"] = self.inputs
        if self.transformations:
            script["tasks"] = self.transformations
        if self.exports:
            script["export"] = self.exports[-1]

        return yaml.safe_dump(script, sort_keys=False, default_flow_style=False)


class SheetQLCompleter(Completer):
    """
    Context-aware autocompletion provider for the interactive shell.

    Provides suggestions for:
    - SQL Keywords
    - Table names (dynamic)
    - Column names (from ALL tables, to allow 'SELECT col...' workflows)
    """

    def __init__(self, schema_cache: Dict[str, List[str]]):
        self.schema_cache = schema_cache
        self.keywords = [
            "SELECT",
            "FROM",
            "WHERE",
            "GROUP BY",
            "ORDER BY",
            "LIMIT",
            "JOIN",
            "LEFT JOIN",
            "RIGHT JOIN",
            "INNER JOIN",
            "ON",
            "AS",
            "DISTINCT",
            "COUNT",
            "SUM",
            "AVG",
            "MIN",
            "MAX",
            "HAVING",
            "CASE",
            "WHEN",
            "THEN",
            "ELSE",
            "END",
            "AND",
            "OR",
            "NOT",
            "IN",
            "IS NULL",
            "IS NOT NULL",
            "LIKE",
            "ILIKE",
            "CAST",
            "DESCRIBE",
            "SHOW TABLES",
            "PRAGMA",
            "EXPORT",
            "PIVOT",
            "UNION",
            "ALL",
        ]

    def get_completions(self, document, complete_event):
        word = document.get_word_before_cursor(WORD=True)
        upper_text = document.text_before_cursor.upper()

        parts = upper_text.split()
        last_word = ""

        if parts:
            if document.text_before_cursor.endswith(" ") or word == "":
                last_word = parts[-1]
            elif len(parts) > 1:
                last_word = parts[-2]

        tables = list(self.schema_cache.keys())
        suggestions = []

        # Context: Expecting a Table Name
        if last_word in ["FROM", "JOIN", "UPDATE", "INTO", "DESCRIBE"]:
            suggestions.extend([(t, "Table") for t in tables])

        # Context: General SQL (Keywords + Columns)
        else:
            suggestions.extend([(k, "Keyword") for k in self.keywords])
            suggestions.extend([(t, "Table") for t in tables])

            for table_name in tables:
                cols = self.schema_cache.get(table_name, [])
                suggestions.extend([(c, f"Column ({table_name})") for c in cols])

        # Filter and yield matching results
        for suggestion, meta in suggestions:
            if suggestion.lower().startswith(word.lower()):
                yield Completion(
                    suggestion, start_position=-len(word), display_meta=meta
                )


class SheetQL:
    """
    Main application controller. Orchestrates DB, UI, and ETL logic.
    """

    PROMPT_SQL = "SQL> "
    PROMPT_CONTINUE = "  -> "
    DEFAULT_EXPORT_FILENAME = "query_result.xlsx"
    HISTORY_MAX_LEN = 50

    def __init__(self, logger: logging.Logger) -> None:
        self.logger = logger
        self.console = Console()
        self.db_connection: Optional[duckdb.DuckDBPyConnection] = None
        self.results_to_save: Dict[str, pd.DataFrame] = {}
        self.history: deque[str] = deque(maxlen=self.HISTORY_MAX_LEN)
        self.schema_cache: Dict[str, List[str]] = {}
        self.recorder = SessionRecorder()
        self.session = None

        if PROMPT_TOOLKIT_AVAILABLE:
            self.session = PromptSession(history=None)

    def run_interactive(self) -> None:
        """Starts the interactive REPL session."""
        try:
            self._display_welcome()
            self._init_db()

            if initial_paths := self._prompt_for_paths(
                title="Select Data Files",
                filetypes=[
                    (
                        "Supported Files",
                        "*.xlsx *.xls *.csv *.parquet *.json *.jsonl *.ndjson",
                    ),
                    ("All files", "*.*"),
                ],
                allow_multiple=True,
            ):
                self._load_data(initial_paths)
                self.logger.info("[bold green]--- 🦆 DuckDB is ready ---[/bold green]")
                self._list_tables()
                self._run_interactive_loop()

        except Exception as e:
            self.logger.critical(f"Fatal error in interactive loop: {e}", exc_info=True)
        finally:
            self.logger.info("[bold cyan]👋 Goodbye![/bold cyan]")

    def run_batch(self, config_path: str) -> None:
        """Runs a headless batch job from a YAML config."""
        self.logger.info(f"🚀 Batch mode: '{config_path}'")
        if not YAML_AVAILABLE:
            self.logger.error("PyYAML is not installed.")
            return

        try:
            with open(config_path, "r") as f:
                config = yaml.safe_load(f)
        except Exception as e:
            self.logger.error(f"Failed to load config: {e}")
            return

        self._init_db()
        self._execute_yaml_script(config)

    def _init_db(self) -> None:
        """Initializes DuckDB with memory limits."""
        self.db_connection = duckdb.connect(database=":memory:")
        try:
            self.db_connection.execute("SET memory_limit='75%';")
        except Exception:
            self.logger.debug(
                "DuckDB memory limit config failed. Proceeding with defaults."
            )

    def _display_welcome(self) -> None:
        """Prints startup banner."""
        self.console.print("[bold green]--- SheetQL Professional ---[/bold green]")
        self.console.print(
            "Commands: [yellow].help[/yellow], [yellow].load[/yellow], [yellow].dump <file>[/yellow]"
        )

        status = []
        status.append(
            "[green]Rust-Excel[/green]"
            if CALAMINE_AVAILABLE
            else "[red]Rust-Excel[/red]"
        )
        status.append(
            "[green]Stream-Write[/green]"
            if XLSXWRITER_AVAILABLE
            else "[red]Stream-Write[/red]"
        )
        status.append(
            "[green]Autocomplete[/green]"
            if PROMPT_TOOLKIT_AVAILABLE
            else "[red]Autocomplete[/red]"
        )
        self.console.print(f"Engine Status: {', '.join(status)}")

    def _prompt_for_paths(
        self, title: str, filetypes: List[Tuple[str, str]], allow_multiple: bool
    ) -> Optional[List[str]]:
        """Gets file paths via GUI or CLI."""
        if TKINTER_AVAILABLE:
            root = tk.Tk()
            root.withdraw()
            if allow_multiple:
                paths = filedialog.askopenfilenames(title=title, filetypes=filetypes)
            else:
                paths = [filedialog.askopenfilename(title=title, filetypes=filetypes)]
            root.destroy()
            return list(paths) if paths and paths[0] else None

        self.console.print(f"\n[cyan]Enter paths for: {title}[/cyan]")
        paths_input = self.console.input("[bold]Path(s): [/bold]")
        raw_paths = [p.strip().strip("'\"") for p in paths_input.split(",")]
        return [p for p in raw_paths if p and os.path.exists(p)]

    def _load_data(self, file_paths: List[str]) -> List[str]:
        """Loads files using Zero-Copy views or Pandas bridges."""
        if not self.db_connection:
            return []
        loaded_tables = []

        with self.console.status("[bold green]Linking files...[/bold green]"):
            for file_path in file_paths:
                try:
                    clean_path = str(file_path).replace("\\", "/")
                    ext = os.path.splitext(file_path)[1].lower()
                    base = re.sub(
                        r"[^a-zA-Z0-9_]+",
                        "_",
                        os.path.splitext(os.path.basename(file_path))[0],
                    )
                    table_name = ""

                    if ext == ".parquet":
                        table_name = f"{base}_parquet"
                        self.db_connection.execute(
                            f"CREATE OR REPLACE VIEW {table_name} AS SELECT * FROM '{clean_path}'"
                        )
                    elif ext == ".csv":
                        table_name = f"{base}_csv"
                        self.db_connection.execute(
                            f"CREATE OR REPLACE VIEW {table_name} AS SELECT * FROM read_csv_auto('{clean_path}')"
                        )
                    elif ext in [".json", ".jsonl"]:
                        table_name = f"{base}_json"
                        self.db_connection.execute(
                            f"CREATE OR REPLACE VIEW {table_name} AS SELECT * FROM read_json_auto('{clean_path}')"
                        )

                    elif ext in [".xlsx", ".xls"]:
                        engine = "calamine" if CALAMINE_AVAILABLE else None
                        try:
                            context = pd.ExcelFile(file_path, engine=engine)
                        except Exception:
                            context = pd.ExcelFile(file_path)

                        with context as xls:
                            for sheet in xls.sheet_names:
                                df = pd.read_excel(xls, sheet_name=sheet)
                                df.columns = [
                                    re.sub(
                                        r"[^a-zA-Z0-9_]+", "_", str(c).strip()
                                    ).lower()
                                    for c in df.columns
                                ]
                                clean_sheet = re.sub(
                                    r"[^a-zA-Z0-9_]+", "_", sheet
                                ).lower()

                                table_name = f"{base}_{clean_sheet}"
                                self.db_connection.register(table_name, df)
                                loaded_tables.append(table_name)
                                self.recorder.record_load(file_path, table_name)
                        self._update_schema_cache(loaded_tables)
                        continue

                    else:
                        self.logger.warning(f"Skipping unsupported type: {ext}")
                        continue

                    if table_name:
                        loaded_tables.append(table_name)
                        self.recorder.record_load(file_path, table_name)

                except Exception as e:
                    self.logger.error(f"Failed to load '{file_path}': {e}")

        self._update_schema_cache(loaded_tables)
        self.logger.info(f"✔ Loaded {len(loaded_tables)} tables.")
        return loaded_tables

    def _update_schema_cache(self, table_names: List[str]) -> None:
        """Fetches and caches table schemas."""
        if not self.db_connection:
            return
        for table in table_names:
            try:
                schema_df = self.db_connection.execute(f"DESCRIBE {table}").fetchdf()
                self.schema_cache[table] = schema_df["column_name"].tolist()
            except Exception:
                pass

    def _run_interactive_loop(self) -> None:
        """Runs the command input loop."""
        query_buffer = ""
        completer = (
            SheetQLCompleter(self.schema_cache) if PROMPT_TOOLKIT_AVAILABLE else None
        )
        style = Style.from_dict({"prompt": "ansicyan bold"})

        while True:
            prompt_text = self.PROMPT_SQL if not query_buffer else self.PROMPT_CONTINUE
            try:
                if PROMPT_TOOLKIT_AVAILABLE and self.session:
                    line = self.session.prompt(
                        prompt_text,
                        completer=completer,
                        lexer=PygmentsLexer(SqlLexer),
                        style=style,
                    )
                else:
                    line = self.console.input(prompt_text)

                if line.strip().startswith("!"):
                    self._handle_history_rerun(line.strip())
                    query_buffer = ""
                    continue

                query_buffer += line + " "
            except (KeyboardInterrupt, EOFError):
                if self._handle_meta_command(".exit"):
                    break
                query_buffer = ""
                continue

            if line.strip().lower().startswith("."):
                if self._handle_meta_command(line.strip()):
                    break
                query_buffer = ""
                continue

            if query_buffer.strip().endswith(";"):
                query_to_run = query_buffer.strip()
                self.history.append(query_to_run)
                self._execute_query(query_to_run)
                query_buffer = ""

    def _execute_query(self, query: str) -> None:
        """Runs SQL and displays results."""
        if not self.db_connection:
            return
        try:
            with self.console.status("[bold green]Executing...[/bold green]"):
                res = self.db_connection.execute(query).fetchdf()

            if res.empty:
                self.console.print("[yellow]No data returned.[/yellow]")
            else:
                self.logger.info("Query Successful")
                self._display_results_table(res)
                self._prompt_to_stage_results(res, query)
        except Exception as e:
            self.logger.error(f"SQL Error: {e}")

    def _display_results_table(self, df: pd.DataFrame) -> None:
        """Formats DataFrame as a Rich table."""
        table = Table(show_header=True, header_style="bold magenta")
        for col in df.columns:
            table.add_column(str(col))
        for _, row in df.head(15).iterrows():
            table.add_row(*[str(x) for x in row])
        self.console.print(table)
        if len(df) > 15:
            self.console.print(f"... ({len(df)-15} more rows)")

    def _prompt_to_stage_results(self, results: pd.DataFrame, query: str) -> None:
        """Offers to stage results for export."""
        if self.console.input("\nStage for export? (y/n): ").lower().startswith("y"):
            name = self.console.input("Sheet name: ")
            if name:
                self.results_to_save[name] = results
                self.recorder.record_query(name, query)
                self.logger.info(f"Staged '{name}'")

    def _save_to_excel(self, save_path: str) -> None:
        """Writes staged results to Excel."""
        try:
            with self.console.status("[bold green]Saving Excel file...[/bold green]"):
                engine = "xlsxwriter" if XLSXWRITER_AVAILABLE else "openpyxl"
                with pd.ExcelWriter(save_path, engine=engine) as writer:
                    for sheet_name, df in self.results_to_save.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                        if engine == "xlsxwriter":
                            wb = writer.book
                            ws = writer.sheets[sheet_name]
                            header_fmt = wb.add_format(
                                {
                                    "bold": True,
                                    "fg_color": "#4F81BD",
                                    "font_color": "white",
                                }
                            )
                            for col_num, value in enumerate(df.columns.values):
                                ws.write(0, col_num, value, header_fmt)
                            for i, col in enumerate(df.columns):
                                ws.set_column(i, i, 20)
                        else:
                            header_font = Font(bold=True, color="FFFFFF")
                            fill = PatternFill(
                                start_color="4F81BD",
                                end_color="4F81BD",
                                fill_type="solid",
                            )
                            for ws in writer.book.worksheets:
                                for cell in ws[1]:
                                    cell.font = header_font
                                    cell.fill = fill
                                ws.auto_filter.ref = ws.dimensions

            self.logger.info(f"Saved to '{os.path.basename(save_path)}' ({engine})")
            self.recorder.record_export(save_path)
            self.results_to_save.clear()
        except Exception as e:
            self.logger.error(f"Save failed: {e}")

    def _handle_meta_command(self, command_str: str) -> bool:
        """Handles dot-commands."""
        parts = command_str.split()
        cmd = parts[0].lower()

        commands = {
            ".exit": lambda: True,
            ".quit": lambda: True,
            ".help": self._show_help,
            ".tables": self._list_tables,
            ".schema": lambda: self._describe_table(parts),
            ".history": self._show_history,
            ".load": self._add_new_files,
            ".export": self._export_results,
            ".dump": lambda: self._dump_script(parts),
            ".runscript": lambda: self._run_script_interactive(parts),
            ".rename": lambda: self._rename_table(parts),
        }

        if cmd not in commands:
            self.logger.warning(f"Unknown command: {cmd}")
            return False

        should_exit = commands[cmd]()
        if should_exit and cmd in [".exit", ".quit"] and self.results_to_save:
            if (
                self.console.input("Export staged results? (y/n): ")
                .lower()
                .startswith("y")
            ):
                self._export_results()
        return should_exit

    def _dump_script(self, parts: List[str]) -> None:
        """Dumps session to YAML."""
        filename = parts[1] if len(parts) > 1 else "script.yaml"
        try:
            yaml_content = self.recorder.generate_yaml()
            with open(filename, "w") as f:
                f.write(yaml_content)
            self.logger.info(f"Session dumped to '[bold cyan]{filename}[/bold cyan]'")
        except Exception as e:
            self.logger.error(f"Failed to dump script: {e}")

    def _show_help(self) -> None:
        self.console.print("\n[bold]Commands:[/bold]")
        self.console.print(
            "  .help, .tables, .schema <t>, .history, .load, .rename <o> <n>, .export, .exit"
        )
        self.console.print(
            "  [bold yellow].dump <file>[/bold yellow]   Save current session to YAML"
        )
        self.console.print(
            "  [bold yellow].runscript <file>[/bold yellow] Run a YAML script"
        )

    def _list_tables(self) -> None:
        if self.db_connection:
            try:
                tables = self.db_connection.execute("SHOW TABLES").fetchdf()["name"]
                self.console.print(f"\n[cyan]Tables ({len(tables)}):[/cyan]")
                for t in tables:
                    self.console.print(f" - {t}")
            except Exception:
                pass

    def _describe_table(self, parts: List[str]) -> None:
        if len(parts) == 2 and self.db_connection:
            try:
                df = self.db_connection.execute(f"DESCRIBE {parts[1]}").fetchdf()
                t = Table(title=f"Schema: {parts[1]}")
                for c in df.columns:
                    t.add_column(c)
                for _, r in df.iterrows():
                    t.add_row(*[str(x) for x in r])
                self.console.print(t)
            except Exception as e:
                self.logger.error(str(e))

    def _rename_table(self, parts: List[str]) -> None:
        """Renames a table view and updates the schema cache safely."""
        if len(parts) == 3:
            try:
                old_name_input = parts[1]
                new_name = parts[2]

                # Update DB
                self.db_connection.execute(
                    f'ALTER VIEW "{old_name_input}" RENAME TO "{new_name}"'
                )
                self.logger.info(f"Renamed {old_name_input} -> {new_name}")

                # Update Cache (Robust lookup for case-insensitivity)
                actual_key = next(
                    (
                        k
                        for k in self.schema_cache
                        if k.lower() == old_name_input.lower()
                    ),
                    old_name_input,
                )

                if actual_key in self.schema_cache:
                    self.schema_cache[new_name] = self.schema_cache.pop(actual_key)
            except Exception as e:
                self.logger.error(str(e))

    def _show_history(self) -> None:
        for i, c in enumerate(self.history, 1):
            self.console.print(f"{i}: {c}")

    def _handle_history_rerun(self, cmd: str) -> None:
        try:
            idx = int(cmd[1:])
            if 1 <= idx <= len(self.history):
                self._execute_query(self.history[idx - 1])
        except Exception:
            pass

    def _add_new_files(self) -> None:
        if paths := self._prompt_for_paths("Select Files", [("All", "*.*")], True):
            self._load_data(paths)

    def _export_results(self) -> None:
        if not self.results_to_save:
            self.logger.warning("Nothing to export.")
            return

        if path := self._prompt_for_paths(
            "Save Location", [("Excel", "*.xlsx")], False
        ):
            save_dest = path[0] if isinstance(path, list) else path
            self._save_to_excel(save_dest)
        else:
            if not TKINTER_AVAILABLE:
                p = self.console.input("Save Path: ")
                if p:
                    self._save_to_excel(p)

    def _run_script_interactive(self, parts: List[str]) -> None:
        script_path = parts[1] if len(parts) > 1 else None
        if not script_path:
            self.logger.warning("Usage: .runscript <file>")
            return
        if not YAML_AVAILABLE:
            self.logger.error("PyYAML missing.")
            return

        try:
            with open(script_path, "r") as f:
                config = yaml.safe_load(f)
            self._execute_yaml_script(config)
        except Exception as e:
            self.logger.error(f"Script Error: {e}")

    def _execute_yaml_script(self, config: Dict[str, Any]) -> None:
        """Executes operations from YAML config."""
        if "inputs" in config:
            paths = [i["path"] for i in config["inputs"]]
            loaded = self._load_data(paths)
            alias_map = {
                os.path.basename(i["path"]): i.get("alias")
                for i in config["inputs"]
                if i.get("alias")
            }

            for tbl in loaded:
                for fname, alias in alias_map.items():
                    if fname.split(".")[0] in tbl:
                        try:
                            self.db_connection.execute(
                                f'ALTER VIEW "{tbl}" RENAME TO "{alias}"'
                            )
                            self.logger.info(f"Aliased {tbl} -> {alias}")
                        except Exception:
                            pass

        if "tasks" in config:
            for task in config.get("tasks", []):
                try:
                    self.results_to_save[task["name"]] = self.db_connection.execute(
                        task["sql"]
                    ).fetchdf()
                    self.logger.info(f"Task '{task['name']}' complete.")
                except Exception as e:
                    self.logger.error(f"Task '{task['name']}' failed: {e}")

        if "export" in config:
            self._save_to_excel(config["export"]["path"])


def main() -> None:
    parser = argparse.ArgumentParser(description="SheetQL Professional")
    parser.add_argument("-r", "--run", dest="config_path", help="Run batch config")
    parser.add_argument(
        "--debug", action="store_true", help="Enable debug logging to console"
    )
    args = parser.parse_args()

    logger = setup_logging(args.debug)

    tool = SheetQL(logger)
    if args.config_path:
        tool.run_batch(args.config_path)
    else:
        tool.run_interactive()


if __name__ == "__main__":
    main()
