import os
import logging
import time
from collections import deque
from typing import Any, Optional, List, Tuple, Dict

import duckdb
import pandas as pd
from openpyxl.styles import Font, PatternFill
from rich import box
from rich.console import Console, Group
from rich.panel import Panel
from rich.rule import Rule
from rich.table import Table
from rich.tree import Tree

from sheetql.completion import SheetQLCompleter
from sheetql.constants import DEFAULT_MEMORY_LIMIT
from sheetql.duckdb_util import (
    apply_performance_pragmas,
    fetch_columns_by_table,
    quote_duckdb_identifier,
    rename_relation,
)
from sheetql.deps import (
    CALAMINE_AVAILABLE,
    PROMPT_TOOLKIT_AVAILABLE,
    TKINTER_AVAILABLE,
    XLSXWRITER_AVAILABLE,
    YAML_AVAILABLE,
)
from sheetql.session import SessionRecorder
from sheetql.naming import normalize_name
from sheetql.scripting import (
    ScriptConfigError,
    ScriptExport,
    parse_script_config,
    resolve_alias_targets,
)

if PROMPT_TOOLKIT_AVAILABLE:
    from prompt_toolkit import PromptSession
    from prompt_toolkit.auto_suggest import AutoSuggestFromHistory
    from prompt_toolkit.formatted_text import HTML
    from prompt_toolkit.history import InMemoryHistory
    from prompt_toolkit.lexers import PygmentsLexer
    from prompt_toolkit.shortcuts import CompleteStyle
    from prompt_toolkit.styles import Style
    from pygments.lexers.sql import SqlLexer
else:
    PromptSession = None  # type: ignore[assignment]
    InMemoryHistory = None  # type: ignore[assignment]
    PygmentsLexer = None  # type: ignore[assignment]
    SqlLexer = None  # type: ignore[assignment]
    Style = None  # type: ignore[assignment]
    CompleteStyle = None  # type: ignore[assignment]
    HTML = None  # type: ignore[assignment]
    AutoSuggestFromHistory = None  # type: ignore[assignment]

if YAML_AVAILABLE:
    import yaml
else:
    yaml = None  # type: ignore[assignment]

if TKINTER_AVAILABLE:
    import tkinter as tk
    from tkinter import filedialog
else:
    tk = None  # type: ignore[assignment]
    filedialog = None  # type: ignore[assignment]


class SheetQL:
    """Main application controller."""

    PROMPT_SQL = "SQL> "
    PROMPT_CONTINUE = "  -> "
    DEFAULT_EXPORT_FILENAME = "query_result.xlsx"
    HISTORY_MAX_LEN = 50
    INTERACTIVE_PREVIEW_ROWS = 15

    def __init__(self, logger: logging.Logger) -> None:
        self.logger = logger
        self.console = Console(soft_wrap=True, highlight=False)
        self.db_connection: Optional[duckdb.DuckDBPyConnection] = None
        self.results_to_save: Dict[str, pd.DataFrame] = {}
        self.history: deque[str] = deque(maxlen=self.HISTORY_MAX_LEN)
        self.schema_cache: Dict[str, List[str]] = {}
        self.recorder = SessionRecorder()

        self.loaded_files_map: Dict[str, List[str]] = {}

        self.session = None
        if (
            PROMPT_TOOLKIT_AVAILABLE
            and PromptSession is not None
            and InMemoryHistory is not None
            and CompleteStyle is not None
            and HTML is not None
            and AutoSuggestFromHistory is not None
        ):
            try:
                # complete_while_typing MUST stay compatible with enable_history_search:
                # prompt_toolkit disables complete_while_typing whenever enable_history_search
                # is True (see PromptSession._create_default_buffer). We prefer live SQL
                # completions while typing; use ↑/↓ for history instead of Ctrl+R search.
                self.session = PromptSession(
                    history=InMemoryHistory(),
                    enable_history_search=False,
                    complete_style=CompleteStyle.COLUMN,
                    complete_while_typing=True,
                    auto_suggest=AutoSuggestFromHistory(),
                    reserve_space_for_menu=8,
                    bottom_toolbar=lambda: HTML(
                        '<style fg="#a3a3a3">'
                        "SheetQL — "
                        ".help · .tables · .files · .peek · .load · .export · .dump · .exit"
                        "</style>"
                    ),
                )
            except Exception:
                self.session = None

    # --- Lifecycle / entrypoints -------------------------------------------------

    def run_interactive(self) -> None:
        try:
            self._display_welcome()
            self._init_db()

            if initial_paths := self._prompt_for_paths(
                title="Select Data Files",
                filetypes=[
                    (
                        "Supported Files",
                        "*.xlsx *.xls *.csv *.parquet *.json *.jsonl *.ndjson",
                    ),
                    ("All files", "*.*"),
                ],
                allow_multiple=True,
            ):
                self._load_data(initial_paths)
                self.logger.info("[bold green]--- 🦆 DuckDB is ready ---[/bold green]")
                self._list_tables()
                self._run_interactive_loop()

        except Exception as e:
            self.logger.critical(f"Fatal error in interactive loop: {e}", exc_info=True)
        finally:
            self.logger.info("[bold cyan]👋 Goodbye![/bold cyan]")

    def run_batch(self, config_path: str) -> None:
        self.logger.info(f"🚀 Batch mode: '{config_path}'")
        if not YAML_AVAILABLE or yaml is None:
            self.logger.error("PyYAML is not installed.")
            return

        try:
            with open(config_path, "r") as f:
                config = yaml.safe_load(f)
        except Exception as e:
            self.logger.error(f"Failed to load config: {e}")
            return

        self._init_db()
        self._execute_yaml_script(config)

    # --- DB / schema -------------------------------------------------------------

    def _init_db(self) -> None:
        self.db_connection = duckdb.connect(database=":memory:")
        try:
            self.db_connection.execute(f"SET memory_limit='{DEFAULT_MEMORY_LIMIT}';")
        except Exception:
            self.logger.debug("DuckDB memory limit config failed. Using defaults.")
        if self.db_connection:
            apply_performance_pragmas(self.db_connection)

    def _update_schema_cache(self, table_names: List[str]) -> None:
        if not self.db_connection or not table_names:
            return
        for table, cols in fetch_columns_by_table(
            self.db_connection, table_names
        ).items():
            self.schema_cache[table] = cols

    # --- UI helpers --------------------------------------------------------------

    def _display_welcome(self) -> None:
        status = []
        status.append(
            "[green]Rust-Excel[/green]"
            if CALAMINE_AVAILABLE
            else "[red]Rust-Excel[/red]"
        )
        status.append(
            "[green]Stream-Write[/green]"
            if XLSXWRITER_AVAILABLE
            else "[red]Stream-Write[/red]"
        )
        status.append(
            "[green]Autocomplete[/green]"
            if PROMPT_TOOLKIT_AVAILABLE
            else "[red]Autocomplete[/red]"
        )
        body = Group(
            "[bold]Type SQL, end statements with[/] [yellow];[/]",
            "[dim]Quick:[/] [yellow].peek[/] [dim]table[/] [dim][n][/]  ·  [yellow].count[/] [dim]table[/]  ·  [yellow].files[/]",
            "[dim]Meta:[/] [yellow].help[/]  [yellow].tables[/]  [yellow].load[/]  [yellow].dump[/] [dim]<file>[/]",
            f"[dim]Engine:[/] {', '.join(status)}",
        )
        self.console.print(
            Panel.fit(
                body,
                title="[bold green]SheetQL[/]",
                border_style="green",
            )
        )

    def _display_results_table(
        self,
        df: pd.DataFrame,
        *,
        title: Optional[str] = None,
        preview_rows: Optional[int] = None,
        col_max_width: int = 48,
    ) -> None:
        row_count = len(df)
        limit = (
            preview_rows if preview_rows is not None else self.INTERACTIVE_PREVIEW_ROWS
        )
        shown = min(limit, row_count) if row_count else 0
        table = Table(
            show_header=True,
            header_style="bold magenta",
            box=box.ROUNDED,
            show_edge=True,
            title=title,
            expand=False,
        )
        for col in df.columns:
            table.add_column(str(col), overflow="ellipsis", max_width=col_max_width)
        if shown > 0:
            preview = df.head(shown).fillna("").astype(str)
            for row in preview.values:
                table.add_row(*row.tolist())
        caption_parts = [
            f"{row_count:,} row(s)",
            f"{len(df.columns)} column(s)",
        ]
        if row_count > shown:
            caption_parts.append(f"showing first {shown:,}")
        table.caption = " · ".join(caption_parts)
        self.console.print(table)

    # --- Path / file prompts -----------------------------------------------------

    def _prompt_for_paths(
        self, title: str, filetypes: List[Tuple[str, str]], allow_multiple: bool
    ) -> Optional[List[str]]:
        if TKINTER_AVAILABLE and tk is not None and filedialog is not None:
            root = tk.Tk()
            root.withdraw()
            if allow_multiple:
                paths = filedialog.askopenfilenames(title=title, filetypes=filetypes)
            else:
                paths = [filedialog.askopenfilename(title=title, filetypes=filetypes)]
            root.destroy()
            return list(paths) if paths and paths[0] else None

        self.console.print(
            Panel(
                f"[bold]{title}[/]\n"
                "[dim]Comma-separated paths, or one path per line. "
                "Quotes around paths are stripped.[/]",
                title="[cyan]Paths[/]",
                border_style="cyan",
                expand=False,
            )
        )
        paths_input = self.console.input("[bold cyan]Path(s):[/] ")
        raw_paths = [p.strip().strip("'\"") for p in paths_input.split(",")]
        return [p for p in raw_paths if p and os.path.exists(p)]

    def _prompt_for_save_path(self) -> Optional[str]:
        """Prompts the user for a new file save path."""
        if TKINTER_AVAILABLE and tk is not None and filedialog is not None:
            root = tk.Tk()
            root.withdraw()

            root.lift()
            root.attributes("-topmost", True)

            save_path = filedialog.asksaveasfilename(
                title="Select Save Location",
                initialfile=self.DEFAULT_EXPORT_FILENAME,
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
            )
            root.destroy()
            return save_path if save_path else None

        self.console.print(
            Panel(
                "[bold]Export destination[/]\n"
                "[dim]File path for the Excel workbook (.xlsx).[/]",
                title="[cyan]Save[/]",
                border_style="cyan",
                expand=False,
            )
        )
        save_path_input = self.console.input(
            f"[bold cyan]Save path[/] [dim](default: {self.DEFAULT_EXPORT_FILENAME})[/]: "
        )
        if not save_path_input:
            save_path_input = self.DEFAULT_EXPORT_FILENAME

        directory = os.path.dirname(save_path_input)
        if directory and not os.path.exists(directory):
            try:
                os.makedirs(directory)
            except OSError as e:
                self.console.print(
                    f"[red]Error: Could not create directory '{directory}'. {e}[/red]"
                )
                return None
        return save_path_input

    # --- Loading / zero-copy views ----------------------------------------------

    def _escape_sql_path(self, path: str) -> str:
        """Escapes single quotes in file paths to prevent SQL injection/errors."""
        return path.replace("'", "''")

    def _load_data(self, file_paths: List[str]) -> List[str]:
        if not self.db_connection:
            return []
        loaded_tables = []

        with self.console.status("[bold green]Linking files...[/bold green]"):
            for file_path in file_paths:
                try:
                    clean_path = str(file_path).replace("\\", "/")
                    sql_safe_path = self._escape_sql_path(clean_path)

                    ext = os.path.splitext(file_path)[1].lower()
                    raw_base = os.path.splitext(os.path.basename(file_path))[0]
                    base = normalize_name(raw_base)
                    table_name = ""

                    generated_tables = []

                    if ext == ".parquet":
                        table_name = f"{base}_parquet"
                        self.db_connection.execute(
                            f"CREATE OR REPLACE VIEW {quote_duckdb_identifier(table_name)} "
                            f"AS SELECT * FROM '{sql_safe_path}'"
                        )
                        generated_tables.append(table_name)
                    elif ext == ".csv":
                        table_name = f"{base}_csv"
                        self.db_connection.execute(
                            f"CREATE OR REPLACE VIEW {quote_duckdb_identifier(table_name)} "
                            f"AS SELECT * FROM read_csv_auto('{sql_safe_path}')"
                        )
                        generated_tables.append(table_name)
                    elif ext in [".json", ".jsonl", ".ndjson"]:
                        table_name = f"{base}_json"
                        self.db_connection.execute(
                            f"CREATE OR REPLACE VIEW {quote_duckdb_identifier(table_name)} "
                            f"AS SELECT * FROM read_json_auto('{sql_safe_path}')"
                        )
                        generated_tables.append(table_name)

                    elif ext in [".xlsx", ".xls"]:
                        engine = "calamine" if CALAMINE_AVAILABLE else None
                        try:
                            context = pd.ExcelFile(file_path, engine=engine)
                        except Exception:
                            context = pd.ExcelFile(file_path)

                        with context as xls:
                            for sheet in xls.sheet_names:
                                if CALAMINE_AVAILABLE:
                                    df = pd.read_excel(
                                        xls, sheet_name=sheet, engine="calamine"
                                    )
                                elif ext == ".xlsx":
                                    try:
                                        df = pd.read_excel(
                                            xls,
                                            sheet_name=sheet,
                                            engine_kwargs={"read_only": True},
                                        )
                                    except Exception:
                                        df = pd.read_excel(xls, sheet_name=sheet)
                                else:
                                    df = pd.read_excel(xls, sheet_name=sheet)
                                # Use normalize_name() consistently — same logic as filename/alias normalization.
                                df.columns = [
                                    normalize_name(str(c)) for c in df.columns
                                ]
                                clean_sheet = normalize_name(sheet)

                                table_name = f"{base}_{clean_sheet}"
                                self.db_connection.register(table_name, df)

                                loaded_tables.append(table_name)
                                generated_tables.append(table_name)
                                self.recorder.record_load(file_path, table_name)

                        self.loaded_files_map[file_path] = generated_tables
                        # Schema cache is populated by the single call at the end of _load_data.
                        continue

                    else:
                        self.logger.warning(f"Skipping unsupported type: {ext}")
                        continue

                    if generated_tables:
                        loaded_tables.extend(generated_tables)
                        self.loaded_files_map[file_path] = generated_tables
                        for t in generated_tables:
                            self.recorder.record_load(file_path, t)

                except Exception as e:
                    self.logger.error(f"Failed to load '{file_path}': {e}")

        self._update_schema_cache(loaded_tables)
        self.logger.info(f"✔ Loaded {len(loaded_tables)} tables.")
        return loaded_tables

    # --- Interactive loop & meta-commands ---------------------------------------

    def _run_interactive_loop(self) -> None:
        query_buffer = ""
        style = (
            Style.from_dict(
                {
                    "prompt": "ansicyan bold",
                    "completion-menu": "bg:ansiblack ansigray",
                    "completion-menu.completion": "bg:ansiblack ansigray",
                    # prompt_toolkit only accepts a subset of ANSI names; avoid
                    # ansibrightwhite (raises ValueError on some versions).
                    "completion-menu.completion.current": "bg:ansiblue bold #ffffff",
                }
            )
            if Style
            else None
        )

        while True:
            completer = (
                SheetQLCompleter(self.schema_cache)
                if PROMPT_TOOLKIT_AVAILABLE
                else None
            )
            prompt_text = self.PROMPT_SQL if not query_buffer else self.PROMPT_CONTINUE
            try:
                if (
                    PROMPT_TOOLKIT_AVAILABLE
                    and self.session
                    and PygmentsLexer
                    and SqlLexer
                    and style
                ):
                    line = self.session.prompt(
                        prompt_text,
                        completer=completer,
                        lexer=PygmentsLexer(SqlLexer),
                        style=style,
                        # Reinforce flags each line (session persists across prompts).
                        enable_history_search=False,
                        complete_while_typing=True,
                    )
                else:
                    line = self.console.input(prompt_text)

                # Handle history re-run (!N) before touching the buffer.
                if line.strip().startswith("!"):
                    self._handle_history_rerun(line.strip())
                    query_buffer = ""
                    continue

                # Handle meta-commands BEFORE appending to the buffer so that
                # typing `.tables` mid-query does not corrupt the query buffer.
                if line.strip().lower().startswith("."):
                    if self._handle_meta_command(line.strip()):
                        break
                    query_buffer = ""
                    continue

                query_buffer += line + " "
            except (KeyboardInterrupt, EOFError):
                if self._handle_meta_command(".exit"):
                    break
                query_buffer = ""
                continue

            if query_buffer.strip().endswith(";"):
                query_to_run = query_buffer.strip()
                self.history.append(query_to_run)
                try:
                    if (
                        self.session is not None
                        and hasattr(self.session, "history")
                        and hasattr(self.session.history, "append_string")
                    ):
                        self.session.history.append_string(query_to_run)
                except Exception:
                    self.logger.debug(
                        "Could not append SQL to prompt history", exc_info=True
                    )
                self._execute_query(query_to_run, offer_stage=True)
                query_buffer = ""

    def _execute_query(
        self,
        query: str,
        *,
        offer_stage: bool = True,
        preview_rows: Optional[int] = None,
    ) -> None:
        if not self.db_connection:
            return
        t0 = time.perf_counter()
        try:
            with self.console.status("[bold green]Executing...[/bold green]"):
                res = self.db_connection.execute(query).fetchdf()
            elapsed_ms = (time.perf_counter() - t0) * 1000

            if res.empty:
                # DuckDB returns a zero-row frame with a single "Count" column for many
                # DDL / mutation statements; distinguish from a normal empty SELECT result.
                if len(res.columns) == 1 and str(res.columns[0]).lower() == "count":
                    self.console.print("[green]Statement completed.[/green]")
                else:
                    self.console.print("[yellow]No data returned.[/yellow]")
                self.console.print(f"[dim]{elapsed_ms:.1f} ms[/]")
            else:
                self.logger.info("Query Successful")
                kw = {}
                if preview_rows is not None:
                    kw["preview_rows"] = preview_rows
                self._display_results_table(res, **kw)
                self.console.print(f"[dim]{elapsed_ms:.1f} ms[/]")
                if offer_stage:
                    self._prompt_to_stage_results(res, query)
        except Exception as e:
            self.logger.error(f"SQL Error: {e}")

    def _prompt_to_stage_results(self, results: pd.DataFrame, query: str) -> None:
        self.console.print()
        if (
            self.console.input(
                "[bold]Stage this result for .export?[/] [dim](y/n)[/]: "
            )
            .lower()
            .startswith("y")
        ):
            name = self.console.input("[bold cyan]Sheet name:[/] ")
            if name:
                self.results_to_save[name] = results
                self.recorder.record_query(name, query)
                self.logger.info(f"Staged '{name}'")

    # --- Exporting ---------------------------------------------------------------

    def _export_dataframe(
        self, df: pd.DataFrame, export: ScriptExport, default_sheet: str
    ) -> None:
        """
        Export a single dataframe to the destination described by ScriptExport.

        Supports: .xlsx, .csv, .json (by extension). For xlsx, uses `export.sheet` or `default_sheet`.
        """
        # If no path is provided in the script, fall back to interactive selection.
        path = export.path
        if not path:
            path = self._prompt_for_save_path()
            if not path:
                self.logger.warning("Export cancelled (no destination selected).")
                return

        directory = os.path.dirname(path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        lower = path.lower()
        if lower.endswith(".xlsx"):
            sheet_name = export.sheet or default_sheet
            prev = dict(self.results_to_save)
            try:
                self.results_to_save.clear()
                self.results_to_save[sheet_name] = df
                self._save_to_excel(path)
            finally:
                self.results_to_save = prev
            return

        if lower.endswith(".csv"):
            df.to_csv(path, index=False)
            self.logger.info(f"Saved to '{os.path.basename(path)}' (csv)")
            return

        if lower.endswith(".json"):
            df.to_json(path, orient="records")
            self.logger.info(f"Saved to '{os.path.basename(path)}' (json)")
            return

        raise ValueError("Unsupported export extension. Use .xlsx, .csv, or .json.")

    @staticmethod
    def _calc_col_width(
        series: pd.Series, col_name: str, min_w: int = 8, max_w: int = 60
    ) -> int:
        """Calculate column width from actual content, capped between min_w and max_w."""
        try:
            max_data = int(series.astype(str).str.len().max()) if len(series) else 0
        except Exception:
            max_data = 0
        return min(max_w, max(min_w, max_data, len(str(col_name))) + 2)

    def _save_to_excel(self, save_path: str) -> None:
        try:
            with self.console.status("[bold green]Saving Excel file...[/bold green]"):
                engine = "xlsxwriter" if XLSXWRITER_AVAILABLE else "openpyxl"
                with pd.ExcelWriter(save_path, engine=engine) as writer:
                    for sheet_name, df in self.results_to_save.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                        if engine == "xlsxwriter":
                            wb = writer.book
                            ws = writer.sheets[sheet_name]
                            header_fmt = wb.add_format(
                                {
                                    "bold": True,
                                    "fg_color": "#4F81BD",
                                    "font_color": "white",
                                }
                            )
                            for col_num, value in enumerate(df.columns.values):
                                ws.write(0, col_num, value, header_fmt)
                            for i, col in enumerate(df.columns):
                                ws.set_column(i, i, self._calc_col_width(df[col], col))
                            # Dropdown filter across all header columns.
                            ws.autofilter(0, 0, len(df), len(df.columns) - 1)
                        else:
                            # Operate on the specific sheet — not all worksheets — to avoid
                            # re-applying styling to previous sheets on each loop iteration.
                            ws = writer.sheets[sheet_name]
                            header_font = Font(bold=True, color="FFFFFF")
                            fill = PatternFill(
                                start_color="4F81BD",
                                end_color="4F81BD",
                                fill_type="solid",
                            )
                            for cell in ws[1]:
                                cell.font = header_font
                                cell.fill = fill
                            for i, col in enumerate(df.columns):
                                col_letter = ws.cell(row=1, column=i + 1).column_letter
                                ws.column_dimensions[col_letter].width = (
                                    self._calc_col_width(df[col], col)
                                )
                            ws.auto_filter.ref = ws.dimensions

            self.logger.info(f"Saved to '{os.path.basename(save_path)}' ({engine})")
            self.recorder.record_export(save_path)
            self.results_to_save.clear()
        except Exception as e:
            self.logger.error(f"Save failed: {e}")

    def _export_results(self) -> None:
        """Exports staged results using the correct save path."""
        if not self.results_to_save:
            self.logger.warning("Nothing to export.")
            return

        if path := self._prompt_for_save_path():
            self._save_to_excel(path)

    # --- Meta commands & helpers -------------------------------------------------

    def _handle_meta_command(self, command_str: str) -> bool:
        parts = command_str.split()
        cmd = parts[0].lower()

        commands = {
            ".exit": lambda: True,
            ".quit": lambda: True,
            ".help": self._show_help,
            ".tables": self._list_tables,
            ".schema": lambda: self._describe_table(parts),
            ".history": self._show_history,
            ".load": self._add_new_files,
            ".export": self._export_results,
            ".dump": lambda: self._dump_script(parts),
            ".runscript": lambda: self._run_script_interactive(parts),
            ".rename": lambda: self._rename_table(parts),
            ".files": self._list_loaded_files,
            ".peek": lambda: self._peek_table(parts),
            ".count": lambda: self._count_table_rows(parts),
            ".cwd": self._show_cwd,
            ".clear": self._clear_screen,
        }

        if cmd not in commands:
            self.logger.warning(f"Unknown command: {cmd}")
            return False

        should_exit = commands[cmd]()
        if should_exit and cmd in [".exit", ".quit"] and self.results_to_save:
            if (
                self.console.input("[bold]Export staged results?[/] [dim](y/n)[/]: ")
                .lower()
                .startswith("y")
            ):
                self._export_results()
        return should_exit

    def _dump_script(self, parts: List[str]) -> None:
        filename = parts[1] if len(parts) > 1 else "script.yaml"
        try:
            yaml_content = self.recorder.generate_yaml()
            with open(filename, "w") as f:
                f.write(yaml_content)
            self.logger.info(f"Session dumped to '[bold cyan]{filename}[/bold cyan]'")
        except Exception as e:
            self.logger.error(f"Failed to dump script: {e}")

    def _show_help(self) -> None:
        help_text = (
            "[cyan].help[/]              This list\n"
            "[cyan].tables[/]            List DuckDB tables\n"
            "[cyan].files[/]             Files you opened → table names\n"
            "[cyan].peek[/] [dim]t [n][/]         First [dim]n[/] rows of table [dim]t[/] (default 15)\n"
            "[cyan].count[/] [dim]t[/]           Row count for table [dim]t[/]\n"
            "[cyan].schema[/] [dim]t[/]          Describe table [dim]t[/]\n"
            "[cyan].history[/]           Numbered history; [dim]![/][dim]n[/] re-runs a query\n"
            "[cyan].load[/]               Add files (picker or paths here)\n"
            "[cyan].rename[/] [dim]old new[/]   Rename a table / view\n"
            "[cyan].export[/]             Write staged sheets to Excel\n"
            "[cyan].dump[/] [dim][file][/]       Save session as YAML\n"
            "[cyan].runscript[/] [dim]file[/]   Run a YAML batch script\n"
            "[cyan].cwd[/]                Show working directory (for paths)\n"
            "[cyan].clear[/]              Clear the terminal screen\n"
            "[cyan].exit[/]  [cyan].quit[/]        Leave (offers export if staged)"
        )
        self.console.print(
            Panel(
                help_text,
                title="[bold]Meta-commands[/]",
                border_style="yellow",
                expand=False,
            )
        )

    def _list_tables(self) -> None:
        if self.db_connection:
            try:
                tables = self.db_connection.execute("SHOW TABLES").fetchdf()["name"]
                names = tables.tolist() if hasattr(tables, "tolist") else list(tables)
                tree = Tree(f"[bold]Tables[/] [dim]({len(names)})[/]")
                for t in names:
                    tree.add(f"[cyan]{t}[/]")
                self.console.print(tree)
            except Exception:
                pass

    def _list_loaded_files(self) -> None:
        """Show each file path and the DuckDB table names created from it."""
        if not self.loaded_files_map:
            self.console.print(
                "[dim]Nothing loaded in this session yet. Use .load or restart and pick files.[/]"
            )
            return
        tree = Tree("[bold]Loaded files[/]")
        for path, tables in self.loaded_files_map.items():
            branch = tree.add(f"[cyan]{path}[/]")
            for t in tables:
                branch.add(f"[dim]→[/] [white]{t}[/]")
        self.console.print(tree)

    def _peek_table(self, parts: List[str]) -> None:
        """Run SELECT * … LIMIT n without staging prompts (quick look at a table)."""
        if len(parts) < 2:
            self.logger.warning("Usage: .peek <table> [rows]")
            return
        name = parts[1].strip()
        if not name:
            self.logger.warning("Usage: .peek <table> [rows]")
            return
        n = 15
        if len(parts) >= 3:
            try:
                n = max(1, min(int(parts[2]), 50_000))
            except ValueError:
                self.logger.warning("Rows must be an integer.")
                return
        qn = quote_duckdb_identifier(name)
        sql = f"SELECT * FROM {qn} LIMIT {n};"
        self._execute_query(sql, offer_stage=False, preview_rows=min(n, 500))

    def _count_table_rows(self, parts: List[str]) -> None:
        """Print row count for a table (faster than typing COUNT(*))."""
        if len(parts) != 2:
            self.logger.warning("Usage: .count <table>")
            return
        name = parts[1].strip()
        if not name:
            self.logger.warning("Usage: .count <table>")
            return
        if not self.db_connection:
            return
        t0 = time.perf_counter()
        try:
            qn = quote_duckdb_identifier(name)
            df = self.db_connection.execute(
                f"SELECT COUNT(*) AS row_count FROM {qn}"
            ).fetchdf()
            elapsed_ms = (time.perf_counter() - t0) * 1000
            n = int(df.iloc[0]["row_count"])
            self.console.print(
                f"[bold]{name}[/]: [cyan]{n:,}[/] row(s)  [dim]({elapsed_ms:.1f} ms)[/]"
            )
        except Exception as e:
            self.logger.error(str(e))

    def _show_cwd(self) -> None:
        self.console.print(
            Panel(os.getcwd(), title="[bold]Working directory[/]", expand=False)
        )

    def _clear_screen(self) -> None:
        self.console.clear()
        self.console.print("[dim]Screen cleared · type SQL or .help[/]")

    def _describe_table(self, parts: List[str]) -> None:
        if len(parts) == 2 and self.db_connection:
            name = parts[1].strip()
            if not name:
                self.logger.warning("Usage: .schema <table>")
                return
            try:
                df = self.db_connection.execute(
                    f"DESCRIBE {quote_duckdb_identifier(name)}"
                ).fetchdf()
                t = Table(
                    title=f"Schema: {name}",
                    box=box.ROUNDED,
                    show_header=True,
                    header_style="bold magenta",
                )
                for c in df.columns:
                    t.add_column(c)
                for _, r in df.iterrows():
                    t.add_row(*[str(x) for x in r])
                self.console.print(t)
            except Exception as e:
                self.logger.error(str(e))

    def _rename_table(self, parts: List[str]) -> None:
        if len(parts) == 3:
            try:
                old = parts[1]
                new_raw = parts[2]
                new = normalize_name(new_raw)

                rename_relation(self.db_connection, old, new)

                self.logger.info(f"Renamed {old} -> {new}")

                # Update schema cache.
                actual_key = next(
                    (k for k in self.schema_cache if k.lower() == old.lower()), old
                )
                if actual_key in self.schema_cache:
                    self.schema_cache[new] = self.schema_cache.pop(actual_key)

                # Update loaded_files_map so YAML alias resolution stays consistent.
                for file_path, tables in self.loaded_files_map.items():
                    self.loaded_files_map[file_path] = [
                        new if t == old else t for t in tables
                    ]

            except Exception as e:
                self.logger.error(str(e))

    def _show_history(self) -> None:
        if not self.history:
            self.console.print("[dim]No queries in history yet.[/]")
            return
        self.console.print(Rule("[bold]History[/]", style="dim"))
        for i, c in enumerate(self.history, 1):
            snippet = c.strip().replace("\n", " ")
            if len(snippet) > 100:
                snippet = snippet[:97] + "…"
            self.console.print(f"[dim]{i:>3}.[/] [cyan]{snippet}[/]")

    def _handle_history_rerun(self, cmd: str) -> None:
        try:
            idx = int(cmd[1:])
            if 1 <= idx <= len(self.history):
                self._execute_query(self.history[idx - 1])
        except Exception:
            pass

    def _add_new_files(self) -> None:
        if paths := self._prompt_for_paths("Select Files", [("All", "*.*")], True):
            self._load_data(paths)

    # --- YAML scripting ----------------------------------------------------------

    def _run_script_interactive(self, parts: List[str]) -> None:
        script_path = parts[1] if len(parts) > 1 else None
        if not script_path:
            self.logger.warning("Usage: .runscript <file>")
            return
        if not YAML_AVAILABLE or yaml is None:
            self.logger.error("PyYAML missing.")
            return

        try:
            with open(script_path, "r") as f:
                config = yaml.safe_load(f)
            self._execute_yaml_script(config)
        except Exception as e:
            self.logger.error(f"Script Error: {e}")

    def _execute_yaml_script(self, config: Dict[str, Any]) -> None:
        """Execute operations from YAML config with validation and options."""
        if not self.db_connection:
            self._init_db()

        try:
            inputs, tasks, export, options = parse_script_config(config)
        except ScriptConfigError as e:
            self.logger.error(f"Invalid script config: {e}")
            return

        if options.memory_limit:
            try:
                self.db_connection.execute(
                    f"SET memory_limit='{options.memory_limit}';"
                )
            except Exception:
                self.logger.debug("DuckDB memory limit config failed. Using defaults.")

        if inputs:
            self._load_data([i.path for i in inputs])

            for item in inputs:
                if not item.alias:
                    continue

                found_tables = resolve_alias_targets(self.loaded_files_map, item.path)
                if not found_tables:
                    self.logger.warning(
                        f"No loaded tables found for input '{item.path}'."
                    )
                    continue

                alias_safe = normalize_name(item.alias)

                if len(found_tables) == 1:
                    rename_relation(self.db_connection, found_tables[0], alias_safe)
                    self.logger.info(f"Aliased {found_tables[0]} -> {alias_safe}")
                else:
                    for tbl in found_tables:
                        suffix = tbl.split("_", 1)[-1] if "_" in tbl else "sheet"
                        new_name = normalize_name(f"{item.alias}_{suffix}")
                        rename_relation(self.db_connection, tbl, new_name)
                        self.logger.info(f"Aliased {tbl} -> {new_name}")

        for task in tasks:
            try:
                df = self.db_connection.execute(task.sql).fetchdf()
                if task.export:
                    try:
                        self._export_dataframe(df, task.export, default_sheet=task.name)
                    except Exception as e:
                        self.logger.error(f"Export for task '{task.name}' failed: {e}")
                        if options.stop_on_error:
                            break
                else:
                    self.results_to_save[task.name] = df
                self.logger.info(f"Task '{task.name}' complete.")
            except Exception as e:
                self.logger.error(f"Task '{task.name}' failed: {e}")
                if options.stop_on_error:
                    break

        if export:
            self._save_to_excel(export.path)


__all__ = ["SheetQL"]
